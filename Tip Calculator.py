#Please install openpyxl onto your drive if you haven't done so already. This will not run without it.
#To install it, type in "pip install openpyxl" (without the quotes) in your command prompt before running this script.


from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Fill, Alignment, Side
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter

#Initialize variables
wb = Workbook()
ws = wb.active

#Set up title and sheet frameworks
ws.title = "Tip Calculator"
ws.sheet_properties.tabColor = "1072BA"

ws.merge_cells('A1:C1')
ws['A1'] = 'Restaurant Tip Calculator'

ws['A3'] = 'Please enter the Post-Tax Amount:'

ws['A5'] = 'Subtotal:'
ws['A6'] = 'Tax:'

ws['A8'] = 'Tip Due:'
ws['A10'] = 'Total Due:'

#Formula
ws['C5'] = '=C3*.91'
ws['C6'] = '=C3*.09'

ws['C8'] = '=C5*.15'

ws['C10'] = '=C3+C8'


#Format cells
grayFill = PatternFill(fill_type = 'solid',
                         start_color = 'A9A9A9',
                         end_color = 'A9A9A9')

yellowFill = PatternFill(fill_type = 'solid',
                         start_color = 'FFFF00',
                         end_color = 'FFFF00')

orangeFill = PatternFill(fill_type = 'solid',
                         start_color = 'FFA500',
                         end_color = 'FFA500')

borderStyle = Border(top = Side(border_style='thick', color='FF000000'),    
                              right = Side(border_style='thick', color='FF000000'), 
                              bottom = Side(border_style='thick', color='FF000000'),
                              left = Side(border_style='thick', color='FF000000'))



ws['A1'].font = Font(bold = True, size = 20)
ws['A1'].fill = grayFill
ws['A1'].alignment = Alignment(horizontal = 'center')

ws['A1'].border = borderStyle
ws['B1'].border = borderStyle
ws['C1'].border = borderStyle

column = 1
i = get_column_letter(column)
ws.column_dimensions[i].width = 40


ws['A3'].font = Font(bold = True, size = 12)

ws['A5'].font = Font(size = 12)
ws['A6'].font = Font(size = 12)

ws['A8'].font = Font(bold = True, size = 12)
ws['A10'].font = Font(bold = True, size = 12)

ws['C3'].font = Font(size = 12)
ws['C3'].fill = yellowFill
ws['C3'].border = borderStyle
ws['C3'].number_format = '$#,###0.00'

ws['C5'].font = Font(size = 12)
ws['C5'].number_format = '$#,###0.00'

ws['C6'].font = Font(size = 12)
ws['C6'].number_format = '$#,###0.00'


ws['C8'].font = Font(bold = True, size = 12)
ws['C8'].fill = orangeFill
ws['C8'].border = borderStyle
ws['C8'].number_format = '$#,###0.00'

ws['C10'].font = Font(bold = True, size = 12)
ws['C10'].fill = orangeFill
ws['C10'].border = borderStyle
ws['C10'].number_format = '$#,###0.00'





wb.save('Tip Calculator.xlsx')
